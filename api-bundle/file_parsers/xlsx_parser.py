"""
XLSX parser — extracts text from Excel workbooks using openpyxl.
"""

import io
import logging
from .base import BaseParser

logger = logging.getLogger(__name__)


class XlsxParser(BaseParser):
    """Extract cell values from .xlsx files using openpyxl."""

    supported_mime_types = [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ]

    def parse(self, data: bytes, filename: str = "") -> dict:
        try:
            import openpyxl  # openpyxl
        except ImportError:
            raise RuntimeError(
                "openpyxl is not installed. Run: pip install openpyxl"
            )

        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception as exc:
            logger.error(f"XLSX parsing failed for '{filename}': {exc}")
            raise ValueError(f"Failed to parse XLSX: {exc}") from exc

        sheet_names = wb.sheetnames
        all_rows: list[str] = []
        total_rows = 0

        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            all_rows.append(f"[Sheet: {sheet_name}]")
            for row in ws.iter_rows(values_only=True):
                cells = [str(cell) for cell in row if cell is not None]
                if cells:
                    all_rows.append(" | ".join(cells))
                    total_rows += 1

        wb.close()

        full_text = "\n".join(all_rows)
        return {
            "text": full_text,
            "sheet_names": sheet_names,
            "row_count": total_rows,
        }
